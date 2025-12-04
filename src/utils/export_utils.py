"""
导出工具 - 将论文数据导出为不同格式
"""
import os
import csv
import json
import logging
from typing import List, Dict, Any
from datetime import datetime
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

class ExportUtils:
    """导出工具类"""
    
    def __init__(self, output_dir: str = 'data/exports'):
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
    
    def export_to_csv(self, papers: List[Dict[str, Any]], filename: str = None) -> str:
        """导出为CSV格式"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"papers_export_{timestamp}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                if not papers:
                    return filepath
                
                # 定义CSV字段
                fieldnames = [
                    'id', 'external_id', 'title', 'abstract', 'authors', 
                    'journal', 'publication_date', 'doi', 'url', 'source',
                    'predicted_category', 'classification_confidence'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for paper in papers:
                    # 处理作者列表
                    authors_str = '; '.join(paper.get('authors', [])) if isinstance(paper.get('authors'), list) else str(paper.get('authors', ''))
                    
                    row = {
                        'id': paper.get('id', ''),
                        'external_id': paper.get('external_id', ''),
                        'title': paper.get('title', ''),
                        'abstract': paper.get('abstract', ''),
                        'authors': authors_str,
                        'journal': paper.get('journal', ''),
                        'publication_date': paper.get('publication_date', ''),
                        'doi': paper.get('doi', ''),
                        'url': paper.get('url', ''),
                        'source': paper.get('source', ''),
                        'predicted_category': paper.get('predicted_category', ''),
                        'classification_confidence': paper.get('classification_confidence', '')
                    }
                    writer.writerow(row)
            
            self.logger.info(f"CSV导出完成: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"CSV导出失败: {e}")
            raise
    
    def export_to_excel(self, papers: List[Dict[str, Any]], filename: str = None) -> str:
        """导出为Excel格式"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl未安装，无法导出Excel格式")
            
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"papers_export_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "论文数据"
            
            # 设置标题样式
            title_font = Font(bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            
            # 定义列标题
            headers = [
                'ID', '外部ID', '标题', '摘要', '作者', '期刊', 
                '发表日期', 'DOI', 'URL', '数据源', '预测分类', '分类置信度'
            ]
            
            # 写入标题行
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = title_alignment
            
            # 写入数据
            for row, paper in enumerate(papers, 2):
                authors_str = '; '.join(paper.get('authors', [])) if isinstance(paper.get('authors'), list) else str(paper.get('authors', ''))
                
                data = [
                    paper.get('id', ''),
                    paper.get('external_id', ''),
                    paper.get('title', ''),
                    paper.get('abstract', ''),
                    authors_str,
                    paper.get('journal', ''),
                    paper.get('publication_date', ''),
                    paper.get('doi', ''),
                    paper.get('url', ''),
                    paper.get('source', ''),
                    paper.get('predicted_category', ''),
                    paper.get('classification_confidence', '')
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # 调整列宽
            column_widths = [8, 15, 50, 80, 30, 25, 12, 20, 30, 10, 15, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # 添加统计工作表
            if papers:
                self._add_statistics_sheet(wb, papers)
            
            wb.save(filepath)
            self.logger.info(f"Excel导出完成: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}")
            raise
    
    def _add_statistics_sheet(self, workbook: Workbook, papers: List[Dict[str, Any]]):
        """添加统计工作表"""
        ws = workbook.create_sheet("统计信息")
        
        # 基本统计
        ws['A1'] = "基本统计"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "总论文数:"
        ws['B3'] = len(papers)
        
        # 按数据源统计
        source_counts = {}
        category_counts = {}
        
        for paper in papers:
            source = paper.get('source', '未知')
            category = paper.get('predicted_category', '未分类')
            
            source_counts[source] = source_counts.get(source, 0) + 1
            if category:
                category_counts[category] = category_counts.get(category, 0) + 1
        
        # 数据源分布
        ws['A5'] = "数据源分布:"
        ws['A5'].font = Font(bold=True)
        
        row = 6
        for source, count in source_counts.items():
            ws[f'A{row}'] = source
            ws[f'B{row}'] = count
            row += 1
        
        # 分类分布
        ws[f'A{row + 1}'] = "分类分布:"
        ws[f'A{row + 1}'].font = Font(bold=True)
        
        row += 2
        for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = category
            ws[f'B{row}'] = count
            row += 1
    
    def export_to_json(self, papers: List[Dict[str, Any]], filename: str = None) -> str:
        """导出为JSON格式"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"papers_export_{timestamp}.json"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            export_data = {
                'metadata': {
                    'export_time': datetime.now().isoformat(),
                    'total_papers': len(papers),
                    'format_version': '1.0'
                },
                'papers': papers
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"JSON导出完成: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"JSON导出失败: {e}")
            raise
    
    def export_to_pdf(self, papers: List[Dict[str, Any]], filename: str = None) -> str:
        """导出为PDF格式"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab未安装，无法导出PDF格式")
            
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"papers_export_{timestamp}.pdf"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            story.append(Paragraph("医学文献导出报告", title_style))
            story.append(Spacer(1, 12))
            
            # 统计信息
            stats_style = styles['Heading2']
            story.append(Paragraph("统计信息", stats_style))
            
            stats_text = f"""
            <para>
            总论文数: {len(papers)}<br/>
            导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
            </para>
            """
            story.append(Paragraph(stats_text, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # 论文列表
            story.append(Paragraph("论文列表", stats_style))
            story.append(Spacer(1, 12))
            
            for i, paper in enumerate(papers[:50], 1):  # 限制PDF中的论文数量
                paper_style = ParagraphStyle(
                    'PaperStyle',
                    parent=styles['Normal'],
                    fontSize=10,
                    spaceAfter=12
                )
                
                title = paper.get('title', '无标题')
                authors = ', '.join(paper.get('authors', [])[:3])  # 只显示前3个作者
                if len(paper.get('authors', [])) > 3:
                    authors += ' 等'
                
                journal = paper.get('journal', '')
                date = paper.get('publication_date', '')
                
                paper_text = f"""
                <para>
                <b>{i}. {title}</b><br/>
                作者: {authors}<br/>
                期刊: {journal} ({date})<br/>
                分类: {paper.get('predicted_category', '未分类')}<br/>
                </para>
                """
                
                story.append(Paragraph(paper_text, paper_style))
            
            if len(papers) > 50:
                story.append(Paragraph(f"... 还有 {len(papers) - 50} 篇论文未显示", styles['Italic']))
            
            doc.build(story)
            self.logger.info(f"PDF导出完成: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"PDF导出失败: {e}")
            raise
    
    def export_summary_report(self, papers: List[Dict[str, Any]], filename: str = None) -> str:
        """导出摘要报告"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"summary_report_{timestamp}.html"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # 统计分析
            total_papers = len(papers)
            source_stats = {}
            category_stats = {}
            year_stats = {}
            
            for paper in papers:
                # 数据源统计
                source = paper.get('source', '未知')
                source_stats[source] = source_stats.get(source, 0) + 1
                
                # 分类统计
                category = paper.get('predicted_category', '未分类')
                if category:
                    category_stats[category] = category_stats.get(category, 0) + 1
                
                # 年份统计
                date = paper.get('publication_date', '')
                if date and len(date) >= 4:
                    year = date[:4]
                    year_stats[year] = year_stats.get(year, 0) + 1
            
            # 生成HTML报告
            html_content = f"""
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>医学文献摘要报告</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .header {{ text-align: center; margin-bottom: 30px; }}
                    .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
                    .stat-box {{ text-align: center; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }}
                    .chart {{ margin: 20px 0; }}
                    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>医学文献摘要报告</h1>
                    <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                
                <div class="stats">
                    <div class="stat-box">
                        <h3>{total_papers}</h3>
                        <p>总论文数</p>
                    </div>
                    <div class="stat-box">
                        <h3>{len(source_stats)}</h3>
                        <p>数据源数</p>
                    </div>
                    <div class="stat-box">
                        <h3>{len(category_stats)}</h3>
                        <p>分类数</p>
                    </div>
                </div>
                
                <h2>数据源分布</h2>
                <table>
                    <tr><th>数据源</th><th>论文数</th><th>占比</th></tr>
            """
            
            for source, count in sorted(source_stats.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_papers) * 100
                html_content += f"<tr><td>{source}</td><td>{count}</td><td>{percentage:.1f}%</td></tr>"
            
            html_content += """
                </table>
                
                <h2>分类分布</h2>
                <table>
                    <tr><th>分类</th><th>论文数</th><th>占比</th></tr>
            """
            
            for category, count in sorted(category_stats.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_papers) * 100
                html_content += f"<tr><td>{category}</td><td>{count}</td><td>{percentage:.1f}%</td></tr>"
            
            html_content += """
                </table>
                
                <h2>年份分布</h2>
                <table>
                    <tr><th>年份</th><th>论文数</th></tr>
            """
            
            for year, count in sorted(year_stats.items(), reverse=True):
                html_content += f"<tr><td>{year}</td><td>{count}</td></tr>"
            
            html_content += """
                </table>
            </body>
            </html>
            """
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"摘要报告导出完成: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"摘要报告导出失败: {e}")
            raise